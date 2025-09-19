#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ordena CFDI 2024 en la hoja "ingresos" de un Excel tipo "cedula 2024".

- Lee XMLs en un directorio (recursivo)
- Extrae campos clave de CFDI 3.3 / 4.0 (Emisor, Receptor, Subtotal, IVA, Total, Fecha, UUID)
- Filtra "ingresos" por RFC emisor si se proporciona (opcional). Si no, incluye todo.
- Genera una tabla con todas las facturas ordenadas por fecha
- Calcula totales por mes (Subtotal, IVA, Total) y los inserta como tabla de resumen
- Escribe o reemplaza la hoja "ingresos" en el Excel de salida

Uso:
  python3 scripts/cfdi_to_cedula.py \
    --xml-dir "$HOME/Desktop/sat/facturas/cfdi2024" \
    --excel   "$HOME/Desktop/SAT/cedula 2024.xlsx" \
    --sheet   ingresos \
    [--rfc    XAXX010101000]

Requisitos: pandas, lxml, openpyxl
  pip install pandas lxml openpyxl
"""

import argparse
import os
import sys
import json
import re
import zipfile
import tempfile
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional

import pandas as pd
from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

CFDI_NS = {
    'cfdi': 'http://www.sat.gob.mx/cfd/3',
    'cfdi40': 'http://www.sat.gob.mx/cfd/4',
    'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
}


def find_first(elem: etree._Element, xpaths: List[str]) -> Optional[etree._Element]:
    for xp in xpaths:
        res = elem.xpath(xp, namespaces=CFDI_NS)
        if res:
            return res[0]
    return None


def get_attr(elem: Optional[etree._Element], attr: str) -> Optional[str]:
    if elem is None:
        return None
    return elem.get(attr)


def parse_cfdi(xml_path: Path) -> Optional[Dict[str, Any]]:
    try:
        with xml_path.open('rb') as f:
            tree = etree.parse(f)
        root = tree.getroot()
        # Comprobante puede ser cfdi:Comprobante (3.3) o {http://www.sat.gob.mx/cfd/4}Comprobante (4.0)
        comp = find_first(root, ['//cfdi:Comprobante', '//cfdi40:Comprobante'])
        if comp is None:
            return None
        emisor = find_first(comp, ['./cfdi:Emisor', './cfdi40:Emisor'])
        receptor = find_first(comp, ['./cfdi:Receptor', './cfdi40:Receptor'])
        imp = find_first(comp, ['./cfdi:Impuestos', './cfdi40:Impuestos'])
        tfd = find_first(comp, ['./cfdi:Complemento/tfd:TimbreFiscalDigital', './cfdi40:Complemento/tfd:TimbreFiscalDigital'])

        # Montos
        subtotal = comp.get('SubTotal') or comp.get('Subtotal') or '0'
        total = comp.get('Total') or '0'
        # IVA: buscar Traslado IVA 002
        iva = '0'
        if imp is not None:
            traslados = imp.xpath('./cfdi:Traslados/cfdi:Traslado | ./cfdi40:Traslados/cfdi40:Traslado', namespaces=CFDI_NS)
            iva_sum = 0.0
            for t in traslados:
                impto = t.get('Impuesto')
                if impto in ('002', 'IVA', '002.0'):
                    iva_val = t.get('Importe') or '0'
                    try:
                        iva_sum += float(iva_val)
                    except Exception:
                        pass
            iva = f"{iva_sum:.2f}"

        fecha = comp.get('Fecha') or comp.get('fecha')
        # UUID
        uuid = get_attr(tfd, 'UUID')

        data = {
            'Fecha': fecha,
            'Serie': comp.get('Serie') or comp.get('serie'),
            'Folio': comp.get('Folio') or comp.get('folio'),
            'UUID': uuid,
            'RFC Emisor': get_attr(emisor, 'Rfc') or get_attr(emisor, 'RFC') or get_attr(emisor, 'rfc'),
            'Nombre Emisor': get_attr(emisor, 'Nombre') or get_attr(emisor, 'nombre'),
            'RFC Receptor': get_attr(receptor, 'Rfc') or get_attr(receptor, 'RFC') or get_attr(receptor, 'rfc'),
            'Nombre Receptor': get_attr(receptor, 'Nombre') or get_attr(receptor, 'nombre'),
            'Subtotal': float(subtotal or 0),
            'IVA': float(iva or 0),
            'Total': float(total or 0),
            'Moneda': comp.get('Moneda') or 'MXN',
            'MetodoPago': comp.get('MetodoPago') or comp.get('MetodoDePago') or '',
            'FormaPago': comp.get('FormaPago') or '',
            'UsoCFDI': get_attr(receptor, 'UsoCFDI') or '',
            'XML': str(xml_path),
        }
        # Normalizar fecha
        if data['Fecha']:
            try:
                # Formatos típicos CFDI: 2024-08-15T12:34:56
                dt = datetime.fromisoformat(data['Fecha'].replace('Z', ''))
                data['Fecha'] = dt
                data['Año'] = dt.year
                data['Mes'] = dt.month
            except Exception:
                data['Año'] = None
                data['Mes'] = None
        return data
    except Exception as e:
        print(f"[WARN] Error parseando {xml_path}: {e}")
        return None


def extract_zips_and_scan_xml_files(xml_dir: Path) -> List[Path]:
    """Extraer archivos ZIP y escanear directorio en busca de archivos XML."""
    xml_files = []
    zip_files_to_delete = []
    
    # Buscar en carpetas ingresos y gastos
    search_dirs = [xml_dir]
    ingresos_dir = xml_dir / 'ingresos'
    gastos_dir = xml_dir / 'gastos'
    
    if ingresos_dir.exists():
        search_dirs.append(ingresos_dir)
    if gastos_dir.exists():
        search_dirs.append(gastos_dir)
    
    for search_dir in search_dirs:
        print(f"[INFO] Buscando archivos en: {search_dir}")
        
        # Extraer archivos ZIP primero
        for zip_pattern in ['*.zip', '*.ZIP']:
            for zip_file in search_dir.rglob(zip_pattern):
                print(f"[INFO] Extrayendo ZIP: {zip_file}")
                try:
                    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                        # Crear directorio temporal para extraer
                        extract_dir = zip_file.parent / f"temp_extract_{zip_file.stem}"
                        extract_dir.mkdir(exist_ok=True)
                        zip_ref.extractall(extract_dir)
                        
                        # Buscar XMLs en el directorio extraído
                        for xml_pattern in ['*.xml', '*.XML']:
                            xml_files.extend(extract_dir.rglob(xml_pattern))
                    
                    zip_files_to_delete.append(zip_file)
                except Exception as e:
                    print(f"[WARN] Error extrayendo {zip_file}: {e}")
        
        # Buscar XMLs directos
        for xml_pattern in ['*.xml', '*.XML']:
            xml_files.extend(search_dir.rglob(xml_pattern))
    
    # Eliminar archivos ZIP después de extraer
    for zip_file in zip_files_to_delete:
        try:
            zip_file.unlink()
            print(f"[INFO] ZIP eliminado: {zip_file}")
        except Exception as e:
            print(f"[WARN] Error eliminando {zip_file}: {e}")
    
    return xml_files


def load_user_rfc_from_runtime_config(project_root: Path) -> Optional[str]:
    try:
        runtime = project_root / 'server' / 'uploads' / 'runtime-config.json'
        if runtime.exists():
            cfg = json.loads(runtime.read_text())
            rfc = (cfg.get('userRfc') or '').strip().upper()
            return rfc or None
    except Exception:
        pass
    return None


def main():
    ap = argparse.ArgumentParser(description='Ordenar CFDI a Excel hoja ingresos con totales por mes')
    ap.add_argument('--xml-dir', required=True, help='Directorio raíz con XML de CFDI (recursivo)')
    ap.add_argument('--excel', required=True, help='Ruta del Excel de salida (e.g. "cedula 2024.xlsx")')
    ap.add_argument('--sheet', default='ingresos', help='Nombre de hoja destino (default: ingresos)')
    ap.add_argument('--rfc', default=None, help='RFC propio para filtrar ingresos (opcional)')
    args = ap.parse_args()

    xml_dir = Path(os.path.expanduser(args.xml_dir)).resolve()
    excel_path = Path(os.path.expanduser(args.excel)).resolve()
    sheet_name = args.sheet

    if not xml_dir.exists():
        print(f"[ERROR] Directorio XML no existe: {xml_dir}")
        sys.exit(1)

    project_root = Path(__file__).resolve().parents[1]
    user_rfc = (args.rfc or load_user_rfc_from_runtime_config(project_root) or '').upper()

    # Recolectar XMLs
    xml_files = extract_zips_and_scan_xml_files(xml_dir)
    if not xml_files:
        print(f"[WARN] No se encontraron XML en: {xml_dir}")

    rows: List[Dict[str, Any]] = []
    for p in xml_files:
        rec = parse_cfdi(p)
        if not rec:
            continue
        if user_rfc:
            # Consideramos ingresos cuando el emisor es el RFC propio
            if (rec.get('RFC Emisor') or '').upper() != user_rfc:
                continue
        rows.append(rec)

    if not rows:
        print("[INFO] No hay registros (revisa RFC o XMLs). Se generará hoja vacía con encabezados.")

    df = pd.DataFrame(rows)
    if not df.empty:
        # Ordenar por fecha
        df = df.sort_values(by=['Fecha', 'Serie', 'Folio'], na_position='last')
        # Totales por mes
        df['Mes'] = df['Mes'].fillna(0).astype(int)
        monthly = df.groupby('Mes', dropna=False).agg({
            'Subtotal': 'sum',
            'IVA': 'sum',
            'Total': 'sum'
        }).reset_index().sort_values('Mes')
    else:
        monthly = pd.DataFrame({'Mes': [], 'Subtotal': [], 'IVA': [], 'Total': []})

    # Escribir a Excel (reemplazando hoja destino)
    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        # Si no existe, crear nuevo libro
        from openpyxl import Workbook
        wb = Workbook()
        # borrar hoja por defecto
        default_sheet = wb.active
        wb.remove(default_sheet)

    # Si existe la hoja, eliminarla para re-crear limpia
    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        wb.remove(ws_old)
    ws = wb.create_sheet(sheet_name)

    # Cargar hoja 'gastos' para deducciones por mes (opcional)
    gastos_by_month = {m: 0.0 for m in range(1, 13)}
    gastos_iva_by_month = {m: 0.0 for m in range(1, 13)}
    gastos_total_by_month = {m: 0.0 for m in range(1, 13)}
    try:
        if 'gastos' in wb.sheetnames:
            df_g = pd.read_excel(excel_path, sheet_name='gastos')
            # Normalizar nombres de columnas
            cols_norm = {c: str(c).strip().lower() for c in df_g.columns}
            df_g.columns = [cols_norm[c] for c in df_g.columns]
            # Identificar columna de subtotal
            subtotal_col = None
            for c in df_g.columns:
                if c in ('subtotal', 'importe', 'monto', 'total'):
                    subtotal_col = c
                    break
            # Detectar IVA y Total si existen explícitamente
            iva_col = None
            for c in df_g.columns:
                if c in ('iva', 'impuesto', 'impuestos'):
                    iva_col = c
                    break
            total_col = None
            for c in df_g.columns:
                if c == 'total':
                    total_col = c
                    break
            # Identificar mes: por columna 'mes' o derivado de 'fecha'
            mes_series = None
            if 'mes' in df_g.columns:
                mes_series = pd.to_numeric(df_g['mes'], errors='coerce')
            elif 'fecha' in df_g.columns:
                mes_series = pd.to_datetime(df_g['fecha'], errors='coerce').dt.month
            if subtotal_col is not None and mes_series is not None:
                df_g['_mes_'] = mes_series
                tmp_sub = (
                    df_g.dropna(subset=['_mes_'])[['_mes_', subtotal_col]]
                    .groupby('_mes_').sum(numeric_only=True).reset_index()
                )
                for _, r in tmp_sub.iterrows():
                    m = int(r['_mes_']) if pd.notna(r['_mes_']) else None
                    if m and 1 <= m <= 12:
                        try:
                            gastos_by_month[m] = float(r[subtotal_col])
                        except Exception:
                            pass
                if iva_col is not None:
                    tmp_iva = (
                        df_g.dropna(subset=['_mes_'])[['_mes_', iva_col]]
                        .groupby('_mes_').sum(numeric_only=True).reset_index()
                    )
                    for _, r in tmp_iva.iterrows():
                        m = int(r['_mes_']) if pd.notna(r['_mes_']) else None
                        if m and 1 <= m <= 12:
                            try:
                                gastos_iva_by_month[m] = float(r[iva_col])
                            except Exception:
                                pass
                if total_col is not None:
                    tmp_tot = (
                        df_g.dropna(subset=['_mes_'])[['_mes_', total_col]]
                        .groupby('_mes_').sum(numeric_only=True).reset_index()
                    )
                    for _, r in tmp_tot.iterrows():
                        m = int(r['_mes_']) if pd.notna(r['_mes_']) else None
                        if m and 1 <= m <= 12:
                            try:
                                gastos_total_by_month[m] = float(r[total_col])
                            except Exception:
                                pass
    except Exception as e:
        print(f"[WARN] No se pudieron leer deducciones de hoja 'gastos': {e}")

    # Cargar 'tablas isr' para límite inferior por mes (opcional)
    # Estructura: per mes, lista de brackets ordenados por límite inferior ascendente
    isr_brackets: dict[int, list[dict[str, float]]] = {m: [] for m in range(1, 13)}
    try:
        if 'tablas isr' in wb.sheetnames:
            df_isr = pd.read_excel(excel_path, sheet_name='tablas isr', header=1)  # Los encabezados están en fila 2
            # Normalizar columnas
            cols_norm = {c: str(c).strip().lower() for c in df_isr.columns}
            df_isr.columns = [cols_norm[c] for c in df_isr.columns]
            # Detectar columnas
            mes_col = None
            for c in df_isr.columns:
                if c in ('mes', 'month'):
                    mes_col = c
                    break
            lim_col = None
            for c in df_isr.columns:
                if ('lim' in c and 'inf' in c) or 'limite inferior' in c:
                    lim_col = c
                    break
            lim_sup_col = None
            for c in df_isr.columns:
                if ('lim' in c and 'sup' in c) or 'limite superior' in c:
                    lim_sup_col = c
                    break
            tasa_col = None
            for c in df_isr.columns:
                if '%' in c or 'tasa' in c or 'porcentaje' in c:
                    tasa_col = c
                    break
            cuota_col = None
            for c in df_isr.columns:
                if 'cuota' in c:
                    cuota_col = c
                    break

            if lim_col is not None:
                # Si no hay columna de mes, asumimos que aplica a todos los meses
                if mes_col is None:
                    df_isr['_mes_'] = 0
                    mes_col = '_mes_'

                for _, r in df_isr.iterrows():
                    try:
                        m_raw = pd.to_numeric(r[mes_col], errors='coerce')
                        m = int(m_raw) if pd.notna(m_raw) and int(m_raw) in range(1,13) else None
                        if m is None:
                            # Si no especifica mes, replicar a todos
                            meses = range(1,13)
                        else:
                            meses = [m]
                        lim_val = float(pd.to_numeric(r[lim_col], errors='coerce') or 0.0)
                        lim_sup_val = float(pd.to_numeric(r[lim_sup_col], errors='coerce') or 999999999.0) if lim_sup_col and pd.notna(r[lim_sup_col]) else 999999999.0
                        tasa_val_raw = pd.to_numeric(r[tasa_col], errors='coerce') if tasa_col in df_isr.columns else None
                        # soportar 30 o 0.30
                        tasa_val = float(tasa_val_raw) if pd.notna(tasa_val_raw) else 0.0
                        if tasa_val > 1.0:
                            tasa_val = tasa_val / 100.0
                        cuota_fija_raw = pd.to_numeric(r[cuota_col], errors='coerce') if cuota_col in df_isr.columns else None
                        cuota_fija = float(cuota_fija_raw) if pd.notna(cuota_fija_raw) else 0.0
                        for mm in meses:
                            isr_brackets[mm].append({
                                'lim_inferior': lim_val,
                                'lim_superior': lim_sup_val,
                                'tasa': tasa_val,
                                'cuota_fija': cuota_fija,
                            })
                    except Exception:
                        pass

                # Ordenar por límite inferior
                for mm in isr_brackets:
                    isr_brackets[mm].sort(key=lambda d: d['lim_inferior'])
    except Exception as e:
        print(f"[WARN] No se pudo procesar 'tablas isr': {e}")

    # Resumen mensual en primeras filas
    ws.append(["Resumen mensual ingresos 2024"])
    ws.append(["Mes", "Subtotal", "IVA", "Total"])
    for _, r in monthly.iterrows():
        ws.append([
            int(r['Mes']) if pd.notna(r['Mes']) and r['Mes'] != '' else '',
            float(r['Subtotal']) if pd.notna(r['Subtotal']) else 0.0,
            float(r['IVA']) if pd.notna(r['IVA']) else 0.0,
            float(r['Total']) if pd.notna(r['Total']) else 0.0,
        ])

    # Sección CÉDULA DE ISR (según especificación del usuario)
    ws.append([])
    ws.append(["CEDULA DE ISR (resumen)"])
    ws.append([
        "Mes",
        "ingresos gravados periodo",
        "ingresos acumulados", 
        "deducciones del periodo",
        "deducciones acumuladas",
        "base gravable",
        "menos lim inferior",
        "% aplicable",
        "igual dif",
        "impuesto marginal",
        "cuota fija",
        "ISR del ejercicio",
        "ISR a pagar",
    ])

    # Construir dataframe de ISR con acumulados y deducciones
    if not monthly.empty:
        isr_df = monthly.copy()
        isr_df['ingresos_gravados_periodo'] = isr_df['Subtotal'].astype(float)
        isr_df['ingresos_acumulados'] = isr_df['ingresos_gravados_periodo'].cumsum()
        # Deducciones del periodo desde gastos_by_month
        isr_df['deducciones_del_periodo'] = isr_df['Mes'].apply(lambda m: gastos_by_month.get(int(m), 0.0))
        isr_df['deducciones_acumuladas'] = isr_df['deducciones_del_periodo'].cumsum()
        isr_df['base_gravable'] = isr_df['ingresos_gravados_periodo'] - isr_df['deducciones_del_periodo']
        # Calcular ISR según especificaciones exactas del usuario
        menos_list = []
        tasa_list = []
        cuota_list = []
        dif_list = []
        imp_marg_list = []
        isr_ejercicio_list = []
        isr_pagar_list = []
        
        for _, rr in isr_df.iterrows():
            mes_i = int(rr['Mes'])
            base = float(rr['base_gravable'])  # ingresos del mes - deducciones del periodo
            brackets = isr_brackets.get(mes_i, [])
            
            # Encontrar bracket correcto según base gravable
            lim_inf = 0.0
            tasa_ap = 0.0
            cuota_fija = 0.0
            if brackets:
                chosen = None
                for b in brackets:
                    if b['lim_inferior'] <= base <= b['lim_superior']:
                        chosen = b
                        break
                if chosen is None:
                    # Si supera todos los rangos, usar el último
                    chosen = brackets[-1] if brackets else {'lim_inferior': 0.0, 'lim_superior': 999999999.0, 'tasa': 0.0, 'cuota_fija': 0.0}
                lim_inf = chosen['lim_inferior']
                tasa_ap = chosen.get('tasa', 0.0)
                cuota_fija = chosen.get('cuota_fija', 0.0)
            
            # Cálculos según especificaciones del usuario:
            # menos_lim_inferior = base_gravable - limite_inferior
            menos_lim_inf = base - lim_inf
            
            # igual_dif = base_gravable - menos_limite_inferior = base_gravable - (base_gravable - limite_inferior) = limite_inferior
            # Pero según la lógica del usuario: igual_dif = base_gravable - limite_inferior
            igual_dif = base - lim_inf
            
            # impuesto_marginal = igual_dif * tasa
            imp_marg = max(igual_dif, 0.0) * tasa_ap
            
            # ISR del ejercicio = impuesto_marginal + cuota_fija
            isr_ejercicio = imp_marg + cuota_fija
            
            # ISR a pagar = ISR del ejercicio
            isr_a_pagar = isr_ejercicio

            menos_list.append(menos_lim_inf)
            tasa_list.append(tasa_ap)
            cuota_list.append(cuota_fija)
            dif_list.append(igual_dif)
            imp_marg_list.append(imp_marg)
            isr_ejercicio_list.append(isr_ejercicio)
            isr_pagar_list.append(isr_a_pagar)

        isr_df['menos_lim_inferior'] = menos_list
        isr_df['tasa_aplicable'] = tasa_list
        isr_df['igual_dif'] = dif_list
        isr_df['impuesto_marginal'] = imp_marg_list
        isr_df['cuota_fija'] = cuota_list
        isr_df['isr_ejercicio'] = isr_ejercicio_list
        isr_df['isr_a_pagar'] = isr_pagar_list

        for _, r in isr_df.iterrows():
            ws.append([
                int(r['Mes']),
                float(r['ingresos_gravados_periodo']),
                float(r['ingresos_acumulados']),
                float(r['deducciones_del_periodo']),
                float(r['deducciones_acumuladas']),
                float(r['base_gravable']),
                float(r['menos_lim_inferior']),
                float(r['tasa_aplicable']),
                float(r['igual_dif']),
                float(r['impuesto_marginal']),
                float(r['cuota_fija']),
                float(r['isr_ejercicio']),
                float(r['isr_a_pagar']),
            ])
    else:
        ws.append(["(sin datos)"])

    # Línea en blanco
    ws.append([])
    ws.append(["Detalle de facturas (ingresos)"])

    # Detalle
    cols = [
        'Fecha','Serie','Folio','UUID','RFC Emisor','Nombre Emisor',
        'RFC Receptor','Nombre Receptor','Subtotal','IVA','Total',
        'Moneda','MetodoPago','FormaPago','UsoCFDI','XML'
    ]
    # Asegurar columnas
    for c in cols:
        if c not in df.columns:
            df[c] = None
    df = df[cols]

    # Convertir Fecha a string legible
    if not df.empty:
        df['Fecha'] = df['Fecha'].apply(lambda d: d.strftime('%Y-%m-%d %H:%M:%S') if isinstance(d, datetime) else (str(d) if d is not None else ''))

    # Marcar inicio de tabla de detalle para crear tabla estructurada
    detail_header_row = ws.max_row + 1
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    detail_last_row = ws.max_row
    if detail_last_row >= detail_header_row:
        first_col = 1
        last_col = len(cols)
        ref = f"{get_column_letter(first_col)}{detail_header_row}:{get_column_letter(last_col)}{detail_last_row}"
        try:
            tbl = Table(displayName="TablaIngresos", ref=ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tbl.tableStyleInfo = style
            ws.add_table(tbl)
        except Exception as e:
            print(f"[WARN] No se pudo crear tabla estructurada: {e}")

    # Hoja 'totales' con totales de ingresos y gastos por mes
    sheet_totales = 'totales'
    if sheet_totales in wb.sheetnames:
        wb.remove(wb[sheet_totales])
    ws_tot = wb.create_sheet(sheet_totales)

    # Totales Ingresos
    ws_tot.append(["Totales Ingresos 2024"])
    ws_tot.append(["Mes", "Subtotal", "IVA", "Total"])
    if not monthly.empty:
        for _, r in monthly.iterrows():
            ws_tot.append([
                int(r['Mes']) if pd.notna(r['Mes']) and r['Mes'] != '' else '',
                float(r['Subtotal']) if pd.notna(r['Subtotal']) else 0.0,
                float(r['IVA']) if pd.notna(r['IVA']) else 0.0,
                float(r['Total']) if pd.notna(r['Total']) else 0.0,
            ])
    # Crear tabla ingresos totales
    tot_ing_header = 1 + 1  # título en fila 1, encabezado en fila 2
    tot_ing_last = ws_tot.max_row
    if tot_ing_last >= (tot_ing_header + 1):
        ref = f"A{tot_ing_header}:D{tot_ing_last}"
        try:
            tbl2 = Table(displayName="TotalesIngresos", ref=ref)
            tbl2.tableStyleInfo = TableStyleInfo(name="TableStyleLight11", showRowStripes=True)
            ws_tot.add_table(tbl2)
        except Exception as e:
            print(f"[WARN] No se pudo crear tabla de totales ingresos: {e}")

    # Espacio y Totales Gastos
    ws_tot.append([])
    ws_tot.append(["Totales Gastos 2024"])
    ws_tot.append(["Mes", "Subtotal", "IVA", "Total"])
    start_gastos_header = ws_tot.max_row
    for m in range(1, 13):
        ws_tot.append([
            m,
            gastos_by_month.get(m, 0.0),
            gastos_iva_by_month.get(m, 0.0),
            gastos_total_by_month.get(m, 0.0) or (gastos_by_month.get(m, 0.0) + gastos_iva_by_month.get(m, 0.0)),
        ])
    end_gastos = ws_tot.max_row
    # Crear tabla gastos totales
    if end_gastos >= (start_gastos_header + 1):
        ref = f"A{start_gastos_header}:D{end_gastos}"
        try:
            tbl3 = Table(displayName="TotalesGastos", ref=ref)
            tbl3.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
            ws_tot.add_table(tbl3)
        except Exception as e:
            print(f"[WARN] No se pudo crear tabla de totales gastos: {e}")

    wb.save(excel_path)
    print(f"[OK] Hoja '{sheet_name}' escrita en: {excel_path}")
    print(f"[OK] Registros: {len(df)} | Meses en resumen: {len(monthly)}")


if __name__ == '__main__':
    main()
