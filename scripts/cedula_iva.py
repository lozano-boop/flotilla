#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generar Cédula del IVA (resumen) a partir de datos procesados de CFDI.

Calcula:
- IVA Causado (por ventas/ingresos)
- IVA Acreditable (por compras/gastos) 
- IVA Retenido
- IVA a Pagar o Saldo a Favor

Uso:
  python3 scripts/cedula_iva.py --excel "$HOME/Desktop/SAT/cedula 2024.xlsx"
"""

import argparse
import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def main():
    ap = argparse.ArgumentParser(description='Generar Cédula del IVA (resumen)')
    ap.add_argument('--excel', required=True, help='Ruta del Excel con datos procesados')
    ap.add_argument('--sheet', default='cedula_iva', help='Nombre de hoja destino (default: cedula_iva)')
    args = ap.parse_args()

    excel_path = Path(os.path.expanduser(args.excel)).resolve()
    sheet_name = args.sheet

    if not excel_path.exists():
        print(f"[ERROR] Archivo Excel no existe: {excel_path}")
        return

    # Cargar workbook
    wb = load_workbook(excel_path)

    # Leer datos de ingresos (IVA Causado)
    iva_causado_by_month = {m: 0.0 for m in range(1, 13)}
    try:
        if 'INGRESOS' in wb.sheetnames:
            df_ingresos = pd.read_excel(excel_path, sheet_name='INGRESOS')
            
            # Los datos están directamente al inicio: Mes, Subtotal, IVA, Total
            if len(df_ingresos) > 0 and 'Mes' in str(df_ingresos.iloc[0, 0]):
                # Leer directamente las primeras filas
                for i in range(1, min(13, len(df_ingresos))):
                    try:
                        mes = int(df_ingresos.iloc[i, 0]) if pd.notna(df_ingresos.iloc[i, 0]) else None
                        iva_val = float(df_ingresos.iloc[i, 2]) if pd.notna(df_ingresos.iloc[i, 2]) else 0.0
                        if mes and 1 <= mes <= 12:
                            iva_causado_by_month[mes] = iva_val
                    except Exception:
                        pass
    except Exception as e:
        print(f"[WARN] Error leyendo IVA causado de INGRESOS: {e}")

    # Leer datos de gastos (IVA Acreditable)
    iva_acreditable_by_month = {m: 0.0 for m in range(1, 13)}
    try:
        if 'gastos' in wb.sheetnames:
            df_gastos = pd.read_excel(excel_path, sheet_name='gastos')
            # Normalizar columnas
            cols_norm = {c: str(c).strip().lower() for c in df_gastos.columns}
            df_gastos.columns = [cols_norm[c] for c in df_gastos.columns]
            
            # Buscar columna de IVA
            iva_col = None
            for c in df_gastos.columns:
                if c in ('iva', 'impuesto', 'impuestos'):
                    iva_col = c
                    break
            
            # Buscar columna de mes
            mes_series = None
            if 'mes' in df_gastos.columns:
                mes_series = pd.to_numeric(df_gastos['mes'], errors='coerce')
            elif 'fecha' in df_gastos.columns:
                mes_series = pd.to_datetime(df_gastos['fecha'], errors='coerce').dt.month
            
            if iva_col and mes_series is not None:
                df_gastos['_mes_'] = mes_series
                tmp_iva = (
                    df_gastos.dropna(subset=['_mes_'])[['_mes_', iva_col]]
                    .groupby('_mes_').sum(numeric_only=True).reset_index()
                )
                for _, r in tmp_iva.iterrows():
                    m = int(r['_mes_']) if pd.notna(r['_mes_']) else None
                    if m and 1 <= m <= 12:
                        try:
                            iva_acreditable_by_month[m] = float(r[iva_col])
                        except Exception:
                            pass
    except Exception as e:
        print(f"[WARN] Error leyendo IVA acreditable de gastos: {e}")

    # Crear hoja de Cédula del IVA
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    # Título y encabezados
    ws.append(["CEDULA DEL IVA 2024 (resumen)"])
    ws.append([])
    ws.append([
        "Mes",
        "IVA Causado",
        "IVA Acreditable", 
        "IVA Retenido",
        "Diferencia",
        "IVA a Pagar",
        "Saldo a Favor"
    ])

    # Datos por mes
    iva_retenido_estimado = 0.0  # Se puede ajustar si tienes datos reales
    
    for mes in range(1, 13):
        iva_causado = iva_causado_by_month.get(mes, 0.0)
        iva_acreditable = iva_acreditable_by_month.get(mes, 0.0)
        iva_retenido = iva_retenido_estimado
        
        # IVA Causado - IVA Acreditable - IVA Retenido
        diferencia = iva_causado - iva_acreditable - iva_retenido
        
        iva_a_pagar = max(diferencia, 0.0)
        saldo_a_favor = abs(min(diferencia, 0.0))
        
        ws.append([
            mes,
            iva_causado,
            iva_acreditable,
            iva_retenido,
            diferencia,
            iva_a_pagar,
            saldo_a_favor
        ])

    # Crear tabla
    header_row = 3
    last_row = ws.max_row
    if last_row >= header_row + 1:
        ref = f"A{header_row}:G{last_row}"
        try:
            tbl = Table(displayName="CedulaIVA", ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium6", showRowStripes=True)
            ws.add_table(tbl)
        except Exception as e:
            print(f"[WARN] No se pudo crear tabla: {e}")

    # Totales anuales
    ws.append([])
    ws.append(["TOTALES ANUALES"])
    total_causado = sum(iva_causado_by_month.values())
    total_acreditable = sum(iva_acreditable_by_month.values())
    total_diferencia = total_causado - total_acreditable - (iva_retenido_estimado * 12)
    
    ws.append(["Total IVA Causado", total_causado])
    ws.append(["Total IVA Acreditable", total_acreditable])
    ws.append(["Total IVA Retenido", iva_retenido_estimado * 12])
    ws.append(["Diferencia Anual", total_diferencia])
    ws.append(["IVA a Pagar Anual", max(total_diferencia, 0.0)])
    ws.append(["Saldo a Favor Anual", abs(min(total_diferencia, 0.0))])

    wb.save(excel_path)
    print(f"[OK] Cédula del IVA creada en hoja '{sheet_name}': {excel_path}")
    print(f"[OK] IVA Causado anual: ${total_causado:,.2f}")
    print(f"[OK] IVA Acreditable anual: ${total_acreditable:,.2f}")
    print(f"[OK] Diferencia anual: ${total_diferencia:,.2f}")

if __name__ == '__main__':
    main()
