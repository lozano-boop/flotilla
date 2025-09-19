#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generar Cédula Unificada (ISR + IVA) en una sola hoja.
Incluye detección de duplicados por UUID y cálculo correcto de IVA acreditable.

Uso:
  python3 scripts/cedula_unificada.py --xml-dir "$HOME/Desktop/SAT/FACTURAS/CFDI2024" --excel "$HOME/Desktop/SAT/cedula 2024.xlsx"
"""

import argparse
import os
import sys
import json
import re
import zipfile
import tempfile
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional, Set

import pandas as pd
from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

CFDI_NS = {
    'cfdi': 'http://www.sat.gob.mx/cfd/3',
    'cfdi40': 'http://www.sat.gob.mx/cfd/4',
    'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
}

def find_first(elem: etree._Element, xpaths: List[str]) -> Optional[etree._Element]:
    for xp in xpaths:
        res = elem.xpath(xp, namespaces=CFDI_NS)
        if res:
            return res[0]
    return None

def get_attr(elem: Optional[etree._Element], attr: str) -> Optional[str]:
    if elem is None:
        return None
    return elem.get(attr)

def parse_cfdi(xml_path: Path) -> Optional[Dict[str, Any]]:
    try:
        with xml_path.open('rb') as f:
            tree = etree.parse(f)
        root = tree.getroot()
        comp = find_first(root, ['//cfdi:Comprobante', '//cfdi40:Comprobante'])
        if comp is None:
            return None
        emisor = find_first(comp, ['./cfdi:Emisor', './cfdi40:Emisor'])
        receptor = find_first(comp, ['./cfdi:Receptor', './cfdi40:Receptor'])
        imp = find_first(comp, ['./cfdi:Impuestos', './cfdi40:Impuestos'])
        tfd = find_first(comp, ['./cfdi:Complemento/tfd:TimbreFiscalDigital', './cfdi40:Complemento/tfd:TimbreFiscalDigital'])

        # Montos
        subtotal = comp.get('SubTotal') or comp.get('Subtotal') or '0'
        total = comp.get('Total') or '0'
        # IVA: buscar Traslado IVA 002
        iva = '0'
        if imp is not None:
            traslados = imp.xpath('./cfdi:Traslados/cfdi:Traslado | ./cfdi40:Traslados/cfdi40:Traslado', namespaces=CFDI_NS)
            iva_sum = 0.0
            for t in traslados:
                impto = t.get('Impuesto')
                if impto in ('002', 'IVA', '002.0'):
                    iva_val = t.get('Importe') or '0'
                    try:
                        iva_sum += float(iva_val)
                    except Exception:
                        pass
            iva = f"{iva_sum:.2f}"

        fecha = comp.get('Fecha') or comp.get('fecha')
        uuid = get_attr(tfd, 'UUID')

        data = {
            'Fecha': fecha,
            'Serie': comp.get('Serie') or comp.get('serie'),
            'Folio': comp.get('Folio') or comp.get('folio'),
            'UUID': uuid,
            'RFC Emisor': get_attr(emisor, 'Rfc') or get_attr(emisor, 'RFC') or get_attr(emisor, 'rfc'),
            'Nombre Emisor': get_attr(emisor, 'Nombre') or get_attr(emisor, 'nombre'),
            'RFC Receptor': get_attr(receptor, 'Rfc') or get_attr(receptor, 'RFC') or get_attr(receptor, 'rfc'),
            'Nombre Receptor': get_attr(receptor, 'Nombre') or get_attr(receptor, 'nombre'),
            'Subtotal': float(subtotal or 0),
            'IVA': float(iva or 0),
            'Total': float(total or 0),
            'Moneda': comp.get('Moneda') or 'MXN',
            'MetodoPago': comp.get('MetodoPago') or comp.get('MetodoDePago') or '',
            'FormaPago': comp.get('FormaPago') or '',
            'UsoCFDI': get_attr(receptor, 'UsoCFDI') or '',
            'XML': str(xml_path),
        }
        # Normalizar fecha
        if data['Fecha']:
            try:
                dt = datetime.fromisoformat(data['Fecha'].replace('Z', ''))
                data['Fecha'] = dt
                data['Año'] = dt.year
                data['Mes'] = dt.month
            except Exception:
                data['Año'] = None
                data['Mes'] = None
        return data
    except Exception as e:
        print(f"[WARN] Error parseando {xml_path}: {e}")
        return None

def extract_zips_and_scan_xml_files(xml_dir: Path) -> List[Path]:
    """Extraer archivos ZIP y escanear directorio en busca de archivos XML."""
    xml_files = []
    zip_files_to_delete = []
    temp_dirs_to_clean = []
    
    # Buscar en carpetas ingresos y gastos
    search_dirs = []
    ingresos_dir = xml_dir / 'ingresos'
    gastos_dir = xml_dir / 'gastos'
    
    if ingresos_dir.exists():
        search_dirs.append(ingresos_dir)
    if gastos_dir.exists():
        search_dirs.append(gastos_dir)
    
    # Si no existen subdirectorios, usar el directorio raíz
    if not search_dirs:
        search_dirs.append(xml_dir)
    
    for search_dir in search_dirs:
        print(f"[INFO] Buscando archivos en: {search_dir}")
        
        # Extraer archivos ZIP primero
        for zip_pattern in ['*.zip', '*.ZIP']:
            for zip_file in search_dir.rglob(zip_pattern):
                print(f"[INFO] Extrayendo ZIP: {zip_file}")
                try:
                    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                        extract_dir = zip_file.parent / f"temp_extract_{zip_file.stem}"
                        extract_dir.mkdir(exist_ok=True)
                        zip_ref.extractall(extract_dir)
                        temp_dirs_to_clean.append(extract_dir)
                        
                        for xml_pattern in ['*.xml', '*.XML']:
                            extracted_xmls = list(extract_dir.rglob(xml_pattern))
                            xml_files.extend(extracted_xmls)
                    
                    zip_files_to_delete.append(zip_file)
                except Exception as e:
                    print(f"[WARN] Error extrayendo {zip_file}: {e}")
        
        # Buscar XMLs directos (excluyendo directorios temporales)
        for xml_pattern in ['*.xml', '*.XML']:
            for xml_file in search_dir.rglob(xml_pattern):
                # Excluir archivos en directorios temporales
                if not any("temp_extract_" in str(xml_file) for temp_dir in temp_dirs_to_clean):
                    xml_files.append(xml_file)
    
    # Eliminar archivos ZIP después de extraer
    for zip_file in zip_files_to_delete:
        try:
            zip_file.unlink()
            print(f"[INFO] ZIP eliminado: {zip_file}")
        except Exception as e:
            print(f"[WARN] Error eliminando {zip_file}: {e}")
    
    # Limpiar directorios temporales
    for temp_dir in temp_dirs_to_clean:
        try:
            shutil.rmtree(temp_dir)
            print(f"[INFO] Directorio temporal eliminado: {temp_dir}")
        except Exception as e:
            print(f"[WARN] Error eliminando directorio temporal {temp_dir}: {e}")
    
    return xml_files

def remove_duplicates_by_uuid(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Eliminar duplicados basándose en UUID."""
    seen_uuids: Set[str] = set()
    unique_rows = []
    duplicates_count = 0
    
    for row in rows:
        uuid = row.get('UUID')
        if uuid and uuid in seen_uuids:
            duplicates_count += 1
            continue
        if uuid:
            seen_uuids.add(uuid)
        unique_rows.append(row)
    
    if duplicates_count > 0:
        print(f"[INFO] Eliminados {duplicates_count} registros duplicados por UUID")
    
    return unique_rows

def classify_by_directory(rows: List[Dict[str, Any]], xml_dir: Path, user_rfc: str) -> tuple:
    """Clasificar registros en ingresos y gastos según directorio y RFC."""
    ingresos = []
    gastos = []
    
    ingresos_dir = xml_dir / 'ingresos'
    gastos_dir = xml_dir / 'gastos'
    
    for row in rows:
        xml_path = Path(row['XML'])
        
        # Clasificar por directorio primero
        if ingresos_dir.exists() and ingresos_dir in xml_path.parents:
            ingresos.append(row)
        elif gastos_dir.exists() and gastos_dir in xml_path.parents:
            gastos.append(row)
        else:
            # Clasificar por RFC si no está en subdirectorios específicos
            if user_rfc and row.get('RFC Emisor') == user_rfc:
                ingresos.append(row)
            else:
                gastos.append(row)
    
    return ingresos, gastos

def main():
    ap = argparse.ArgumentParser(description='Generar Cédula Unificada ISR + IVA')
    ap.add_argument('--xml-dir', required=True, help='Directorio raíz con XML de CFDI')
    ap.add_argument('--excel', required=True, help='Ruta del Excel de salida')
    ap.add_argument('--sheet', default='cedula_unificada', help='Nombre de hoja destino')
    ap.add_argument('--rfc', default=None, help='RFC propio para filtrar ingresos')
    args = ap.parse_args()

    xml_dir = Path(os.path.expanduser(args.xml_dir)).resolve()
    excel_path = Path(os.path.expanduser(args.excel)).resolve()
    sheet_name = args.sheet

    if not xml_dir.exists():
        print(f"[ERROR] Directorio XML no existe: {xml_dir}")
        sys.exit(1)

    user_rfc = (args.rfc or '').upper()

    # Recolectar XMLs
    xml_files = extract_zips_and_scan_xml_files(xml_dir)
    if not xml_files:
        print(f"[WARN] No se encontraron XML en: {xml_dir}")

    # Procesar XMLs
    rows: List[Dict[str, Any]] = []
    for p in xml_files:
        rec = parse_cfdi(p)
        if rec:
            rows.append(rec)

    print(f"[INFO] Registros antes de eliminar duplicados: {len(rows)}")
    
    # Eliminar duplicados por UUID
    rows = remove_duplicates_by_uuid(rows)
    print(f"[INFO] Registros después de eliminar duplicados: {len(rows)}")

    # Clasificar en ingresos y gastos
    ingresos, gastos = classify_by_directory(rows, xml_dir, user_rfc)
    print(f"[INFO] Ingresos: {len(ingresos)}, Gastos: {len(gastos)}")

    # Crear DataFrames
    df_ingresos = pd.DataFrame(ingresos) if ingresos else pd.DataFrame()
    df_gastos = pd.DataFrame(gastos) if gastos else pd.DataFrame()

    # Calcular totales mensuales
    def calculate_monthly_totals(df, name):
        if df.empty:
            return pd.DataFrame()
        
        # Filtrar solo registros válidos con año y mes
        df_valid = df.dropna(subset=['Año', 'Mes'])
        df_2024 = df_valid[df_valid['Año'] == 2024]
        
        print(f"[DEBUG] {name} - Total registros: {len(df)}")
        print(f"[DEBUG] {name} - Registros válidos: {len(df_valid)}")
        print(f"[DEBUG] {name} - Registros 2024: {len(df_2024)}")
        
        if df_2024.empty:
            print(f"[WARN] {name} - No hay datos válidos para 2024")
            return pd.DataFrame()
        
        monthly = (
            df_2024.groupby('Mes')[['Subtotal', 'IVA', 'Total']]
            .sum()
            .reset_index()
        )
        print(f"[INFO] {name} - Meses con datos: {sorted(monthly['Mes'].tolist())}")
        return monthly

    monthly_ingresos = calculate_monthly_totals(df_ingresos, "Ingresos")
    monthly_gastos = calculate_monthly_totals(df_gastos, "Gastos")

    # Cargar/crear workbook
    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    # Eliminar hoja si existe
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    # Cargar tablas ISR
    isr_brackets: dict[int, list[dict[str, float]]] = {m: [] for m in range(1, 13)}
    try:
        if 'tablas isr' in wb.sheetnames:
            df_isr = pd.read_excel(excel_path, sheet_name='tablas isr', header=1)
            cols_norm = {c: str(c).strip().lower() for c in df_isr.columns}
            df_isr.columns = [cols_norm[c] for c in df_isr.columns]
            
            for _, r in df_isr.iterrows():
                try:
                    m_raw = pd.to_numeric(r['mes'], errors='coerce')
                    m = int(m_raw) if pd.notna(m_raw) and int(m_raw) in range(1,13) else None
                    if m is None:
                        meses = range(1,13)
                    else:
                        meses = [m]
                    
                    lim_val = float(pd.to_numeric(r['limite_inferior'], errors='coerce') or 0.0)
                    lim_sup_val = float(pd.to_numeric(r['limite_superior'], errors='coerce') or 999999999.0)
                    tasa_val_raw = pd.to_numeric(r['tasa'], errors='coerce')
                    tasa_val = float(tasa_val_raw) if pd.notna(tasa_val_raw) else 0.0
                    if tasa_val > 1.0:
                        tasa_val = tasa_val / 100.0
                    cuota_fija = float(pd.to_numeric(r['cuota_fija'], errors='coerce') or 0.0)
                    
                    for mm in meses:
                        isr_brackets[mm].append({
                            'lim_inferior': lim_val,
                            'lim_superior': lim_sup_val,
                            'tasa': tasa_val,
                            'cuota_fija': cuota_fija,
                        })
                except Exception:
                    pass

            for mm in isr_brackets:
                isr_brackets[mm].sort(key=lambda d: d['lim_inferior'])
    except Exception as e:
        print(f"[WARN] No se pudo procesar 'tablas isr': {e}")

    # TÍTULO PRINCIPAL
    ws.append(["CEDULA UNIFICADA ISR + IVA 2024"])
    ws.append([])

    # SECCIÓN 1: CÉDULA DE ISR
    ws.append(["CEDULA DE ISR (resumen)"])
    ws.append([
        "Mes", "Ingresos Gravados", "Ingresos Acumulados", "Deducciones Periodo",
        "Deducciones Acumuladas", "Base Gravable", "Menos Lím Inferior",
        "% Aplicable", "Igual Dif", "Impuesto Marginal", "Cuota Fija",
        "ISR del Ejercicio", "ISR Retenido", "ISR a Pagar"
    ])

    # Leer retenciones de ISR si existe la hoja
    isr_retenido_by_month = {m: 0.0 for m in range(1, 13)}
    try:
        if 'retenciones' in wb.sheetnames:
            df_retenciones = pd.read_excel(excel_path, sheet_name='retenciones')
            
            # Agrupar retenciones de ISR por mes
            for _, row in df_retenciones.iterrows():
                try:
                    mes_ini = int(row.get('Mes Inicial', 0)) if pd.notna(row.get('Mes Inicial')) else 0
                    isr_ret = float(row.get('ISR Retenido', 0)) if pd.notna(row.get('ISR Retenido')) else 0.0
                    if 1 <= mes_ini <= 12:
                        isr_retenido_by_month[mes_ini] += isr_ret
                except Exception:
                    pass
            
            print(f"[INFO] ISR retenido cargado por mes: {sum(isr_retenido_by_month.values()):,.2f}")
    except Exception as e:
        print(f"[WARN] Error leyendo retenciones para ISR: {e}")

    # Calcular ISR por mes
    if not monthly_ingresos.empty:
        # Crear diccionarios de gastos por mes
        gastos_by_month = {m: 0.0 for m in range(1, 13)}
        if not monthly_gastos.empty:
            for _, r in monthly_gastos.iterrows():
                mes = int(r['Mes'])
                gastos_by_month[mes] = float(r['Subtotal'])

        isr_df = monthly_ingresos.copy()
        isr_df['ingresos_gravados_periodo'] = isr_df['Subtotal'].astype(float)
        isr_df['ingresos_acumulados'] = isr_df['ingresos_gravados_periodo'].cumsum()
        isr_df['deducciones_del_periodo'] = isr_df['Mes'].apply(lambda m: gastos_by_month.get(int(m), 0.0))
        isr_df['deducciones_acumuladas'] = isr_df['deducciones_del_periodo'].cumsum()
        isr_df['base_gravable'] = isr_df['ingresos_gravados_periodo'] - isr_df['deducciones_del_periodo']

        # Calcular ISR
        for _, rr in isr_df.iterrows():
            mes_i = int(rr['Mes'])
            base = float(rr['base_gravable'])
            brackets = isr_brackets.get(mes_i, [])
            
            lim_inf = 0.0
            tasa_ap = 0.0
            cuota_fija = 0.0
            if brackets:
                chosen = None
                for b in brackets:
                    if b['lim_inferior'] <= base <= b['lim_superior']:
                        chosen = b
                        break
                if chosen is None:
                    chosen = brackets[-1] if brackets else {'lim_inferior': 0.0, 'lim_superior': 999999999.0, 'tasa': 0.0, 'cuota_fija': 0.0}
                lim_inf = chosen['lim_inferior']
                tasa_ap = chosen.get('tasa', 0.0)
                cuota_fija = chosen.get('cuota_fija', 0.0)
            
            menos_lim_inf = base - lim_inf
            igual_dif = base - lim_inf
            imp_marg = max(igual_dif, 0.0) * tasa_ap
            isr_ejercicio = imp_marg + cuota_fija
            
            # ISR Retenido del mes
            isr_retenido_mes = isr_retenido_by_month.get(mes_i, 0.0)
            
            # ISR a Pagar = ISR del Ejercicio - ISR Retenido
            isr_a_pagar = max(isr_ejercicio - isr_retenido_mes, 0.0)

            ws.append([
                mes_i,
                float(rr['ingresos_gravados_periodo']),
                float(rr['ingresos_acumulados']),
                float(rr['deducciones_del_periodo']),
                float(rr['deducciones_acumuladas']),
                float(rr['base_gravable']),
                menos_lim_inf,
                tasa_ap,
                igual_dif,
                imp_marg,
                cuota_fija,
                isr_ejercicio,
                isr_retenido_mes,
                isr_a_pagar
            ])
    else:
        ws.append(["(sin datos de ingresos)"])

    # SECCIÓN 2: CÉDULA DEL IVA
    ws.append([])
    ws.append(["CEDULA DEL IVA (resumen)"])
    ws.append([
        "Mes", "IVA Causado", "IVA Acreditable", "Diferencia", 
        "IVA Retenido", "IVA a Pagar", "Saldo a Favor"
    ])

    # Leer retenciones de IVA si existe la hoja
    iva_retenido_by_month = {m: 0.0 for m in range(1, 13)}
    try:
        if 'retenciones' in wb.sheetnames:
            df_retenciones = pd.read_excel(excel_path, sheet_name='retenciones')
            
            # Agrupar retenciones de IVA por mes
            for _, row in df_retenciones.iterrows():
                try:
                    mes_ini = int(row.get('Mes Inicial', 0)) if pd.notna(row.get('Mes Inicial')) else 0
                    iva_ret = float(row.get('IVA Retenido', 0)) if pd.notna(row.get('IVA Retenido')) else 0.0
                    if 1 <= mes_ini <= 12:
                        iva_retenido_by_month[mes_ini] += iva_ret
                except Exception:
                    pass
            
            print(f"[INFO] IVA retenido cargado por mes: {sum(iva_retenido_by_month.values()):,.2f}")
    except Exception as e:
        print(f"[WARN] Error leyendo retenciones para IVA: {e}")

    # Calcular IVA por mes
    iva_causado_total = 0.0
    iva_acreditable_total = 0.0
    iva_retenido_total = 0.0

    for mes in range(1, 13):
        # IVA Causado (de ingresos)
        iva_causado = 0.0
        if not monthly_ingresos.empty:
            mes_data = monthly_ingresos[monthly_ingresos['Mes'] == mes]
            if not mes_data.empty:
                iva_causado = float(mes_data.iloc[0]['IVA'])

        # IVA Acreditable (de gastos)
        iva_acreditable = 0.0
        if not monthly_gastos.empty:
            mes_data = monthly_gastos[monthly_gastos['Mes'] == mes]
            if not mes_data.empty:
                iva_acreditable = float(mes_data.iloc[0]['IVA'])

        # IVA Retenido (de retenciones)
        iva_retenido = iva_retenido_by_month.get(mes, 0.0)

        # Cálculos según flujo contable
        diferencia = iva_causado - iva_acreditable
        iva_a_pagar = max(diferencia - iva_retenido, 0.0)
        saldo_a_favor = abs(min(diferencia - iva_retenido, 0.0))

        ws.append([
            mes,
            iva_causado,
            iva_acreditable,
            diferencia,
            iva_retenido,
            iva_a_pagar,
            saldo_a_favor
        ])

        iva_causado_total += iva_causado
        iva_acreditable_total += iva_acreditable
        iva_retenido_total += iva_retenido

    # TOTALES ANUALES
    ws.append([])
    ws.append(["TOTALES ANUALES"])
    diferencia_anual = iva_causado_total - iva_acreditable_total
    iva_a_pagar_anual = max(diferencia_anual - iva_retenido_total, 0.0)
    saldo_a_favor_anual = abs(min(diferencia_anual - iva_retenido_total, 0.0))

    ws.append(["Total IVA Causado", iva_causado_total])
    ws.append(["Total IVA Acreditable", iva_acreditable_total])
    ws.append(["Diferencia Anual", diferencia_anual])
    ws.append(["Total IVA Retenido", iva_retenido_total])
    ws.append(["IVA a Pagar Anual", iva_a_pagar_anual])
    ws.append(["Saldo a Favor Anual", saldo_a_favor_anual])

    wb.save(excel_path)
    print(f"[OK] Cédula Unificada creada en hoja '{sheet_name}': {excel_path}")
    print(f"[OK] Registros procesados: {len(rows)} (sin duplicados)")
    print(f"[OK] Ingresos: {len(ingresos)}, Gastos: {len(gastos)}")
    print(f"[OK] IVA Causado anual: ${iva_causado_total:,.2f}")
    print(f"[OK] IVA Acreditable anual: ${iva_acreditable_total:,.2f}")
    print(f"[OK] IVA a Pagar anual: ${iva_a_pagar_anual:,.2f}")

if __name__ == '__main__':
    main()
