#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procesar XMLs de retenciones (Uber, Didi, etc.) y agregarlos al Excel.

Uso:
  python3 scripts/procesar_retenciones.py --retenciones-dir "$HOME/Desktop/SAT/FACTURAS/CFDI2024/retenciones" --excel "$HOME/Desktop/SAT/cedula 2024.xlsx"
"""

import argparse
import os
import sys
import zipfile
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional

import pandas as pd
from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

RETENCIONES_NS = {
    'retenciones': 'http://www.sat.gob.mx/esquemas/retencionpago/2',
    'plataformasTecnologicas': 'http://www.sat.gob.mx/esquemas/retencionpago/1/PlataformasTecnologicas10',
    'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital'
}

def find_first_retention(elem: etree._Element, xpaths: List[str]) -> Optional[etree._Element]:
    for xp in xpaths:
        res = elem.xpath(xp, namespaces=RETENCIONES_NS)
        if res:
            return res[0]
    return None

def get_attr_retention(elem: Optional[etree._Element], attr: str) -> Optional[str]:
    if elem is None:
        return None
    return elem.get(attr)

def parse_retencion_xml(xml_path: Path) -> Optional[Dict[str, Any]]:
    """Parsear XML de retenciones y extraer datos relevantes."""
    try:
        with xml_path.open('rb') as f:
            tree = etree.parse(f)
        root = tree.getroot()
        
        # Verificar que es un XML de retenciones
        if root.tag != '{http://www.sat.gob.mx/esquemas/retencionpago/2}Retenciones':
            return None
        
        # Datos principales
        folio = root.get('FolioInt')
        fecha_exp = root.get('FechaExp')
        cve_retenc = root.get('CveRetenc')
        
        # Emisor
        emisor = find_first_retention(root, ['//retenciones:Emisor'])
        rfc_emisor = get_attr_retention(emisor, 'RfcE')
        nombre_emisor = get_attr_retention(emisor, 'NomDenRazSocE')
        
        # Receptor
        receptor = find_first_retention(root, ['//retenciones:Nacional'])
        rfc_receptor = get_attr_retention(receptor, 'RfcR')
        nombre_receptor = get_attr_retention(receptor, 'NomDenRazSocR')
        
        # Período
        periodo = find_first_retention(root, ['//retenciones:Periodo'])
        mes_ini = get_attr_retention(periodo, 'MesIni')
        mes_fin = get_attr_retention(periodo, 'MesFin')
        ejercicio = get_attr_retention(periodo, 'Ejercicio')
        
        # Totales
        totales = find_first_retention(root, ['//retenciones:Totales'])
        monto_tot_operacion = float(get_attr_retention(totales, 'MontoTotOperacion') or 0)
        monto_tot_grav = float(get_attr_retention(totales, 'MontoTotGrav') or 0)
        monto_tot_ret = float(get_attr_retention(totales, 'MontoTotRet') or 0)
        
        # Impuestos retenidos
        impuestos = root.xpath('//retenciones:ImpRetenidos', namespaces=RETENCIONES_NS)
        isr_retenido = 0.0
        iva_retenido = 0.0
        
        for imp in impuestos:
            impuesto_tipo = imp.get('ImpuestoRet')
            monto_ret = float(imp.get('MontoRet') or 0)
            
            if impuesto_tipo == '001':  # ISR
                isr_retenido += monto_ret
            elif impuesto_tipo == '002':  # IVA
                iva_retenido += monto_ret
        
        # UUID
        tfd = find_first_retention(root, ['//tfd:TimbreFiscalDigital'])
        uuid = get_attr_retention(tfd, 'UUID')
        fecha_timbrado = get_attr_retention(tfd, 'FechaTimbrado')
        
        # Detalles del servicio (si es plataforma tecnológica)
        servicio = find_first_retention(root, ['//plataformasTecnologicas:DetallesDelServicio'])
        fecha_servicio = get_attr_retention(servicio, 'FechaServ')
        precio_sin_iva = float(get_attr_retention(servicio, 'PrecioServSinIVA') or 0)
        tipo_servicio = get_attr_retention(servicio, 'TipoDeServ')
        
        # Normalizar fechas
        fecha_exp_dt = None
        if fecha_exp:
            try:
                fecha_exp_dt = datetime.fromisoformat(fecha_exp.replace('Z', ''))
            except Exception:
                pass
        
        fecha_servicio_dt = None
        if fecha_servicio:
            try:
                fecha_servicio_dt = datetime.fromisoformat(fecha_servicio + 'T00:00:00')
            except Exception:
                pass
        
        data = {
            'Folio': folio,
            'Fecha Expedición': fecha_exp_dt,
            'Fecha Servicio': fecha_servicio_dt,
            'Clave Retención': cve_retenc,
            'RFC Emisor': rfc_emisor,
            'Nombre Emisor': nombre_emisor,
            'RFC Receptor': rfc_receptor,
            'Nombre Receptor': nombre_receptor,
            'Mes Inicial': int(mes_ini) if mes_ini else None,
            'Mes Final': int(mes_fin) if mes_fin else None,
            'Ejercicio': int(ejercicio) if ejercicio else None,
            'Monto Total Operación': monto_tot_operacion,
            'Monto Total Gravado': monto_tot_grav,
            'Monto Total Retenido': monto_tot_ret,
            'ISR Retenido': isr_retenido,
            'IVA Retenido': iva_retenido,
            'Precio Sin IVA': precio_sin_iva,
            'Tipo Servicio': tipo_servicio,
            'UUID': uuid,
            'Fecha Timbrado': fecha_timbrado,
            'XML': str(xml_path),
        }
        
        return data
        
    except Exception as e:
        print(f"[WARN] Error parseando retención {xml_path}: {e}")
        return None

def extract_and_process_retenciones(retenciones_dir: Path) -> List[Dict[str, Any]]:
    """Extraer ZIPs de retenciones y procesar XMLs."""
    xml_files = []
    temp_dirs_to_clean = []
    zip_files_to_delete = []
    
    print(f"[INFO] Procesando retenciones en: {retenciones_dir}")
    
    # Extraer archivos ZIP
    for zip_pattern in ['*.zip', '*.ZIP']:
        for zip_file in retenciones_dir.rglob(zip_pattern):
            print(f"[INFO] Extrayendo ZIP de retención: {zip_file.name}")
            try:
                with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                    extract_dir = zip_file.parent / f"temp_ret_{zip_file.stem}"
                    extract_dir.mkdir(exist_ok=True)
                    zip_ref.extractall(extract_dir)
                    temp_dirs_to_clean.append(extract_dir)
                    
                    for xml_pattern in ['*.xml', '*.XML']:
                        extracted_xmls = list(extract_dir.rglob(xml_pattern))
                        xml_files.extend(extracted_xmls)
                
                zip_files_to_delete.append(zip_file)
            except Exception as e:
                print(f"[WARN] Error extrayendo {zip_file}: {e}")
    
    # Buscar XMLs directos
    for xml_pattern in ['*.xml', '*.XML']:
        for xml_file in retenciones_dir.rglob(xml_pattern):
            if not any("temp_ret_" in str(xml_file) for temp_dir in temp_dirs_to_clean):
                xml_files.append(xml_file)
    
    print(f"[INFO] XMLs de retenciones encontrados: {len(xml_files)}")
    
    # Procesar XMLs
    retenciones_data = []
    for xml_path in xml_files:
        data = parse_retencion_xml(xml_path)
        if data:
            retenciones_data.append(data)
    
    # Limpiar archivos temporales
    for temp_dir in temp_dirs_to_clean:
        try:
            shutil.rmtree(temp_dir)
            print(f"[INFO] Directorio temporal eliminado: {temp_dir}")
        except Exception as e:
            print(f"[WARN] Error eliminando directorio temporal {temp_dir}: {e}")
    
    # Eliminar ZIPs (opcional - comentar si no se desea)
    # for zip_file in zip_files_to_delete:
    #     try:
    #         zip_file.unlink()
    #         print(f"[INFO] ZIP eliminado: {zip_file}")
    #     except Exception as e:
    #         print(f"[WARN] Error eliminando {zip_file}: {e}")
    
    print(f"[INFO] Retenciones procesadas exitosamente: {len(retenciones_data)}")
    return retenciones_data

def main():
    ap = argparse.ArgumentParser(description='Procesar XMLs de retenciones y agregarlos al Excel')
    ap.add_argument('--retenciones-dir', required=True, help='Directorio con XMLs/ZIPs de retenciones')
    ap.add_argument('--excel', required=True, help='Ruta del Excel de salida')
    ap.add_argument('--sheet', default='retenciones', help='Nombre de hoja destino')
    args = ap.parse_args()

    retenciones_dir = Path(os.path.expanduser(args.retenciones_dir)).resolve()
    excel_path = Path(os.path.expanduser(args.excel)).resolve()
    sheet_name = args.sheet

    if not retenciones_dir.exists():
        print(f"[ERROR] Directorio de retenciones no existe: {retenciones_dir}")
        sys.exit(1)

    # Procesar retenciones
    retenciones_data = extract_and_process_retenciones(retenciones_dir)
    
    if not retenciones_data:
        print("[WARN] No se encontraron retenciones válidas")
        return

    # Crear DataFrame
    df_retenciones = pd.DataFrame(retenciones_data)
    
    # Ordenar por fecha de servicio y emisor
    df_retenciones = df_retenciones.sort_values(['Fecha Servicio', 'Nombre Emisor'], na_position='last')

    # Cargar/crear workbook
    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    # Eliminar hoja si existe
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    # Escribir datos
    for r in dataframe_to_rows(df_retenciones, index=False, header=True):
        ws.append(r)

    # Crear tabla
    if len(df_retenciones) > 0:
        table = Table(
            displayName=f"Tabla_{sheet_name}",
            ref=f"A1:{ws.cell(row=len(df_retenciones)+1, column=len(df_retenciones.columns)).coordinate}"
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=True
        )
        ws.add_table(table)

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path)
    
    print(f"[OK] Retenciones agregadas en hoja '{sheet_name}': {excel_path}")
    print(f"[OK] Total retenciones procesadas: {len(retenciones_data)}")
    
    # Mostrar resumen por emisor
    resumen = df_retenciones.groupby('Nombre Emisor').agg({
        'Monto Total Operación': 'sum',
        'ISR Retenido': 'sum',
        'IVA Retenido': 'sum',
        'UUID': 'count'
    }).rename(columns={'UUID': 'Cantidad'})
    
    print(f"\n[RESUMEN POR EMISOR]")
    for emisor, data in resumen.iterrows():
        print(f"  {emisor}:")
        print(f"    Cantidad: {int(data['Cantidad'])} retenciones")
        print(f"    Monto Total: ${data['Monto Total Operación']:,.2f}")
        print(f"    ISR Retenido: ${data['ISR Retenido']:,.2f}")
        print(f"    IVA Retenido: ${data['IVA Retenido']:,.2f}")

if __name__ == '__main__':
    main()
