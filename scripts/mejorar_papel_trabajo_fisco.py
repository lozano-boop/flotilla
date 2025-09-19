#!/usr/bin/env python3
"""
Script para mejorar el papel de trabajo anual con información de Fisco Agenda 2025
Integra requisitos específicos para actividades empresariales y profesionales

Autor: Sistema CFDI
Fecha: 2024
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
import argparse

def mejorar_papel_trabajo_con_fisco_agenda(excel_path: Path):
    """
    Mejora el papel de trabajo existente con información de Fisco Agenda 2025
    """
    
    wb = load_workbook(excel_path)
    
    # Verificar si existe la hoja papel_trabajo_anual
    if 'papel_trabajo_anual' not in wb.sheetnames:
        print("Error: No se encontró la hoja 'papel_trabajo_anual'. Ejecute primero crear_papel_trabajo_anual.py")
        return
    
    ws = wb['papel_trabajo_anual']
    
    # Configurar estilos
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    fill_legal = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
    fill_warning = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
    
    # Encontrar la última fila con contenido
    last_row = ws.max_row
    while last_row > 1 and not any(ws.cell(row=last_row, column=col).value for col in range(1, 6)):
        last_row -= 1
    
    row = last_row + 3
    
    # SECCIÓN NUEVA: FUNDAMENTOS LEGALES FISCO AGENDA 2025
    ws[f'A{row}'] = "VI. FUNDAMENTOS LEGALES - FISCO AGENDA 2025"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = fill_legal
    row += 1
    
    # Subsección ISR - Actividades Empresariales y Profesionales
    ws[f'A{row}'] = "A) ISR - ACTIVIDADES EMPRESARIALES Y PROFESIONALES"
    ws[f'A{row}'].font = header_font
    row += 1
    
    fundamentos_isr = [
        ("Art. 100 LISR", "Ingresos por actividades empresariales y profesionales"),
        ("Art. 101 LISR", "Momento de acumulación de ingresos"),
        ("Art. 103 LISR", "Deducciones autorizadas"),
        ("Art. 106 LISR", "Deducción de inversiones"),
        ("Art. 109 LISR", "Pagos provisionales mensuales"),
        ("Art. 110 LISR", "Obligaciones específicas del régimen"),
        ("Art. 113-E LISR", "Régimen Simplificado de Confianza (RESICO)"),
        ("Art. 150-152 LISR", "Declaración anual personas físicas")
    ]
    
    for articulo, descripcion in fundamentos_isr:
        ws[f'A{row}'] = articulo
        ws[f'B{row}'] = descripcion
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    row += 1
    
    # Subsección IVA
    ws[f'A{row}'] = "B) IVA - PERSONAS FÍSICAS"
    ws[f'A{row}'].font = header_font
    row += 1
    
    fundamentos_iva = [
        ("Art. 1 LIVA", "Actos gravados con IVA"),
        ("Art. 5 LIVA", "IVA acreditable"),
        ("Art. 6 LIVA", "Momento de causación del impuesto"),
        ("Art. 18-H LIVA", "Retención de IVA por plataformas tecnológicas"),
        ("Art. 32 LIVA", "Obligaciones de personas físicas")
    ]
    
    for articulo, descripcion in fundamentos_iva:
        ws[f'A{row}'] = articulo
        ws[f'B{row}'] = descripcion
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    row += 2
    
    # SECCIÓN: OBLIGACIONES ESPECÍFICAS SEGÚN FISCO AGENDA
    ws[f'A{row}'] = "VII. OBLIGACIONES ESPECÍFICAS - ACTIVIDADES PROFESIONALES"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = fill_warning
    row += 1
    
    obligaciones = [
        "1. DECLARACIÓN ANUAL OBLIGATORIA (Art. 150 LISR)",
        "   • Plazo: Abril del año siguiente",
        "   • Incluir todos los ingresos del ejercicio",
        "   • Aplicar deducciones autorizadas",
        "",
        "2. PAGOS PROVISIONALES MENSUALES (Art. 109 LISR)",
        "   • Plazo: Día 17 del mes siguiente",
        "   • Base: Coeficiente de utilidad del ejercicio anterior",
        "   • Retenciones acreditables",
        "",
        "3. CONTABILIDAD ELECTRÓNICA (Art. 110 LISR)",
        "   • Libros diario y mayor",
        "   • Balanza de comprobación mensual",
        "   • Estados financieros anuales",
        "",
        "4. COMPROBANTES FISCALES DIGITALES",
        "   • Emitir CFDI por servicios prestados",
        "   • Conservar comprobantes de gastos",
        "   • Validar vigencia de certificados",
        "",
        "5. RETENCIONES DE PLATAFORMAS TECNOLÓGICAS",
        "   • ISR retenido: Acreditable contra ISR anual",
        "   • IVA retenido: Acreditable contra IVA causado",
        "   • Conservar constancias de retención"
    ]
    
    for obligacion in obligaciones:
        ws[f'A{row}'] = obligacion
        if obligacion.startswith(("1.", "2.", "3.", "4.", "5.")):
            ws[f'A{row}'].font = header_font
        row += 1
    
    row += 2
    
    # SECCIÓN: DEDUCCIONES AUTORIZADAS ESPECÍFICAS
    ws[f'A{row}'] = "VIII. DEDUCCIONES AUTORIZADAS (Art. 103 LISR)"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = fill_legal
    row += 1
    
    deducciones = [
        ("Gastos directamente relacionados", "Con la prestación del servicio profesional"),
        ("Inversiones", "Deducción anual según tabla de por cientos máximos"),
        ("Sueldos y salarios", "Pagados a trabajadores, incluyendo prestaciones"),
        ("Arrendamiento", "De bienes inmuebles destinados a la actividad"),
        ("Servicios profesionales", "Honorarios pagados a terceros"),
        ("Combustibles y lubricantes", "Para vehículos utilizados en la actividad"),
        ("Mantenimiento", "De equipo e instalaciones"),
        ("Seguros", "Relacionados con la actividad empresarial"),
        ("Cuotas patronales", "IMSS, INFONAVIT, SAR"),
        ("Intereses", "Por préstamos destinados a la actividad")
    ]
    
    for concepto, descripcion in deducciones:
        ws[f'A{row}'] = f"• {concepto}:"
        ws[f'B{row}'] = descripcion
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    row += 2
    
    # SECCIÓN: VERIFICACIONES CRÍTICAS
    ws[f'A{row}'] = "IX. VERIFICACIONES CRÍTICAS ANTES DE DECLARAR"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = fill_warning
    row += 1
    
    verificaciones = [
        "☐ Todos los CFDI emitidos están timbrados y vigentes",
        "☐ Comprobantes de gastos cumplen requisitos fiscales",
        "☐ Retenciones de plataformas corresponden al ejercicio 2024",
        "☐ Pagos provisionales están correctamente calculados",
        "☐ Deducciones cumplen con límites y requisitos legales",
        "☐ Contabilidad electrónica está actualizada",
        "☐ No hay inconsistencias entre ingresos declarados y CFDI emitidos",
        "☐ Saldos de IVA están correctamente determinados",
        "☐ Se conservan todos los comprobantes y documentación",
        "☐ Cálculos de ISR anual están verificados"
    ]
    
    for verificacion in verificaciones:
        ws[f'A{row}'] = verificacion
        row += 1
    
    row += 2
    
    # SECCIÓN: FECHAS IMPORTANTES 2025
    ws[f'A{row}'] = "X. FECHAS IMPORTANTES 2025"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    row += 1
    
    fechas = [
        ("30 de abril 2025", "Fecha límite declaración anual 2024"),
        ("17 de cada mes", "Pagos provisionales mensuales 2025"),
        ("Último día del mes", "Envío contabilidad electrónica"),
        ("31 de diciembre", "Cierre del ejercicio fiscal")
    ]
    
    for fecha, descripcion in fechas:
        ws[f'A{row}'] = fecha
        ws[f'B{row}'] = descripcion
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 25
    
    # Guardar archivo
    wb.save(excel_path)
    print(f"[OK] Papel de trabajo mejorado con Fisco Agenda 2025: {excel_path}")

def main():
    parser = argparse.ArgumentParser(description='Mejorar papel de trabajo con Fisco Agenda 2025')
    parser.add_argument('--excel', required=True, help='Ruta al archivo Excel con papel de trabajo')
    
    args = parser.parse_args()
    
    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"Error: No se encontró el archivo {excel_path}")
        return
    
    mejorar_papel_trabajo_con_fisco_agenda(excel_path)

if __name__ == "__main__":
    main()
